from __future__ import annotations

import csv
import json
import math
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


BASE = Path(r"C:\Users\cwb\Desktop\job\01二维公路软件\平整度验证\新算法平整度栗庙路真值验证")
RAW = BASE / "原始平整度结果"
TRUE = {(200, 300): 1.25, (300, 400): 1.27, (400, 500): 1.43, (500, 600): 1.75}
OUT = Path(r"D:\job\工作COD\2025\公路\XRDataProcess\tmp\docs\iri_truth_kb_analysis.json")


def read_coeff(path: Path):
    values = [float(x.strip()) for x in path.read_text(encoding="utf-8-sig").splitlines() if x.strip()]
    n = int(values[0])
    return values[1:1+n], values[1+n:1+2*n], values[1+2*n:1+3*n]


def select_coeff(speed, thresholds, ks, bs):
    for t, k, b in zip(thresholds, ks, bs):
        if speed <= t:
            return k, b
    return ks[-1], bs[-1]


def read_iri(path: Path):
    result = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        parts = line.split()
        if len(parts) >= 2:
            result.append(float(parts[1]))
    return result


def read_speed(path: Path):
    result = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        parts = line.split()
        if len(parts) >= 2:
            result.append(float(parts[1]))
    return result


def fit_ols(samples):
    x = np.array([s["raw_iri"] for s in samples], dtype=float)
    y = np.array([s["truth"] for s in samples], dtype=float)
    design = np.column_stack([x, np.ones_like(x)])
    k, b = np.linalg.lstsq(design, y, rcond=None)[0]
    pred = k * x + b
    return {"k": float(k), "b": float(b), "pred": pred, "rmse": float(np.sqrt(np.mean((pred-y)**2))), "mae": float(np.mean(abs(pred-y))), "r2": float(1-np.sum((pred-y)**2)/np.sum((y-y.mean())**2))}


def fit_huber(samples, max_iter=80):
    x = np.array([s["raw_iri"] for s in samples], dtype=float)
    y = np.array([s["truth"] for s in samples], dtype=float)
    design = np.column_stack([x, np.ones_like(x)])
    weights = np.ones(len(x))
    k = b = 0.0
    for _ in range(max_iter):
        sw = np.sqrt(weights)
        next_k, next_b = np.linalg.lstsq(design*sw[:, None], y*sw, rcond=None)[0]
        residual = y - (next_k*x + next_b)
        scale = max(1e-6, 1.4826*np.median(abs(residual - np.median(residual))))
        cutoff = 1.345*scale
        next_weights = np.minimum(1.0, cutoff/np.maximum(abs(residual), 1e-12))
        if abs(next_k-k)+abs(next_b-b) < 1e-9:
            k, b, weights = next_k, next_b, next_weights
            break
        k, b, weights = next_k, next_b, next_weights
    pred = k*x + b
    return {"k": float(k), "b": float(b), "pred": pred, "weights": weights, "rmse": float(np.sqrt(np.mean((pred-y)**2))), "mae": float(np.mean(abs(pred-y))), "r2": float(1-np.sum((pred-y)**2)/np.sum((y-y.mean())**2))}


def leave_one_project_out(samples, kind):
    results = []
    projects = sorted(set(s["project"] for s in samples))
    for p in projects:
        train = [s for s in samples if s["project"] != p]
        test = [s for s in samples if s["project"] == p]
        model = fit_huber(train) if kind == "huber" else fit_ols(train)
        err = [model["k"]*s["raw_iri"] + model["b"] - s["truth"] for s in test]
        results.extend(err)
    return {"rmse": float(np.sqrt(np.mean(np.square(results)))), "mae": float(np.mean(np.abs(results))), "max_abs": float(np.max(np.abs(results)))}


def main():
    samples = []
    workbooks = [p for p in sorted(RAW.rglob("*.xlsx")) if not p.name.startswith("~$")]
    for xlsx in workbooks:
        project = xlsx.stem.replace("_IRI_100m", "")
        group = project.split("-", 1)[0]
        book = load_workbook(xlsx, data_only=True, read_only=True)
        sheet = book["Sheet1"]
        for row in sheet.iter_rows(min_row=4, values_only=True):
            start, end = row[0], row[1]
            if not isinstance(start, (int, float)) or not isinstance(end, (int, float)):
                continue
            interval = (int(start), int(end))
            if interval not in TRUE:
                continue
            # 本批工程归档中的 IRI_10m.txt 均为空，以用户指定的100m结果表为
            # 唯一IRI来源；该列已经应用旧K/B，按表内实际车速反算基础IRI。
            raw_iri = float(row[3])
            measured_speed = float(row[9])
            samples.append({
                "group": group,
                "project": project,
                "interval": f"{int(start)}-{int(end)}",
                "truth": TRUE[interval],
                "raw_iri": raw_iri,
                "current_iri": raw_iri,
                "xlsx_iri": raw_iri,
                "speed": measured_speed,
                "old_k_mean": 1.0,
                "old_b_mean": 0.0,
                "xlsx_difference": 0.0,
            })
    by_group = {}
    for group in sorted(set(s["group"] for s in samples), key=int):
        group_samples = [s for s in samples if s["group"] == group]
        ols = fit_ols(group_samples)
        huber = fit_huber(group_samples)
        for s, p1, p2, weight in zip(group_samples, ols["pred"], huber["pred"], huber["weights"]):
            s["ols_pred"] = float(p1)
            s["ols_error"] = float(p1-s["truth"])
            s["huber_pred"] = float(p2)
            s["huber_error"] = float(p2-s["truth"])
            s["huber_weight"] = float(weight)
        by_group[group] = {
            "samples": len(group_samples),
            "speed_min": float(min(s["speed"] for s in group_samples)),
            "speed_max": float(max(s["speed"] for s in group_samples)),
            "speed_mean": float(np.mean([s["speed"] for s in group_samples])),
            "current_rmse": float(np.sqrt(np.mean([(s["current_iri"]-s["truth"])**2 for s in group_samples]))),
            "current_mae": float(np.mean([abs(s["current_iri"]-s["truth"]) for s in group_samples])),
            "ols": {k:v for k,v in ols.items() if k != "pred"},
            "huber": {k:(v.tolist() if hasattr(v, "tolist") else v) for k,v in huber.items() if k not in ("pred", "weights")},
            "ols_cv": leave_one_project_out(group_samples, "ols"),
            "huber_cv": leave_one_project_out(group_samples, "huber"),
        }
    result = {"truth": {f"{a}-{b}":v for (a,b),v in TRUE.items()}, "samples": samples, "groups": by_group, "workbooks": len(workbooks)}
    OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result["groups"], ensure_ascii=False, indent=2))
    print("WORKBOOKS", len(workbooks), "SAMPLES", len(samples), "OUT", OUT)


if __name__ == "__main__":
    main()
