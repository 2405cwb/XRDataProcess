from pathlib import Path
from openpyxl import load_workbook

base = Path(r"C:\Users\cwb\Desktop\job\01二维公路软件\平整度验证\新算法平整度栗庙路真值验证\原始平整度结果")
for path in sorted(base.rglob("*.xlsx"))[:2]:
    book = load_workbook(path, data_only=True, read_only=True)
    print("FILE", path.name)
    print("SHEETS", book.sheetnames)
    for sheet in book.worksheets:
        print("SHEET", sheet.title, sheet.max_row, sheet.max_column)
        for row in sheet.iter_rows(min_row=1, max_row=min(12, sheet.max_row), values_only=True):
            print(row)
