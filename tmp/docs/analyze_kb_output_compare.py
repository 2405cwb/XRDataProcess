import json
import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(r"C:\Users\cwb\Desktop\job\01二维公路软件\平整度验证\新算法平整度栗庙路真值验证")
FOLDER_A = BASE / "v2.2.5.9软件出表 乘以k的平均值"
FOLDER_B = BASE / "内部软件  乘k+b"
OUT = Path(r"D:\job\工作COD\2025\公路\XRDataProcess\tmp\docs\kb_output_compare_data.json")


def number(v):
    return float(v) if isinstance(v, (int, float)) else None


def project_id(path):
    m = re.search(r"(?:^|\\)([357]0-\d\d)__", str(path))
    if not m:
        raise ValueError(f"未识别工程编号: {path}")
    return m.group(1)


def read_folder(folder, label):
    rows = []
    files = sorted(folder.rglob("*.xlsx"))
    for file in files:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        pid = project_id(file)
        for r in range(4, ws.max_row + 1):
            start = number(ws.cell(r, 1).value)
            end = number(ws.cell(r, 2).value)
            iri = number(ws.cell(r, 4).value)
            speed = number(ws.cell(r, 10).value)
            if start is None or end is None or iri is None:
                continue
            speed_group = int(pid.split("-")[0])
            rows.append({
                "algorithm": label,
                "project": pid,
                "speedGroup": speed_group,
                "start": start,
                "end": end,
                "speed": speed or speed_group,
                "iri": iri,
                "sourceFile": str(file),
            })
        wb.close()
    return rows, files


def avg(xs):
    return sum(xs) / len(xs) if xs else 0


def stats(rows):
    n = len(rows)
    ds = [x["delta"] for x in rows]
    ars = [abs(x["relative"]) for x in rows]
    return {
        "count": n,
        "meanA": avg([x["iriA"] for x in rows]),
        "meanB": avg([x["iriB"] for x in rows]),
        "meanDelta": avg(ds),
        "meanAbsDelta": avg([abs(x) for x in ds]),
        "meanRelative": avg([x["relative"] for x in rows]),
        "meanAbsRelative": avg(ars),
        "maxAbsDelta": max([abs(x) for x in ds], default=0),
        "maxAbsRelative": max(ars, default=0),
        "higher": sum(1 for x in ds if x > 0),
        "lower": sum(1 for x in ds if x < 0),
        "same": sum(1 for x in ds if x == 0),
    }


rows_a, files_a = read_folder(FOLDER_A, "A")
rows_b, files_b = read_folder(FOLDER_B, "B")
index_a = {(x["project"], x["start"], x["end"]): x for x in rows_a}
index_b = {(x["project"], x["start"], x["end"]): x for x in rows_b}
keys = sorted(set(index_a) & set(index_b), key=lambda k: (int(k[0].split("-")[0]), k[0], k[1]))
detail = []
for key in keys:
    a, b = index_a[key], index_b[key]
    d = b["iri"] - a["iri"]
    detail.append({
        "speedGroup": a["speedGroup"], "project": a["project"], "start": a["start"], "end": a["end"],
        "interval": f"{int(a['start'])}-{int(a['end'])}", "speed": a["speed"], "iriA": a["iri"], "iriB": b["iri"],
        "delta": d, "relative": d / a["iri"] if a["iri"] else None,
    })

by_speed = []
for speed in (30, 50, 70):
    group = [x for x in detail if x["speedGroup"] == speed]
    by_speed.append({"label": f"{speed} km/h", "speed": speed, **stats(group)})

by_segment = []
for start in (0, 100, 200, 300, 400, 500):
    end = start + 100
    group = [x for x in detail if x["start"] == start and x["end"] == end]
    by_segment.append({"label": f"{start}-{end} m", "start": start, "end": end, **stats(group)})
last_group = [x for x in detail if x["start"] >= 600]
if last_group:
    by_segment.append({"label": "600m后末段", "start": 600, "end": None, **stats(last_group)})

stable_by_speed = []
for speed in (30, 50, 70):
    group = [x for x in detail if x["speedGroup"] == speed and x["start"] >= 200 and x["end"] <= 600]
    stable_by_speed.append({"label": f"{speed} km/h", "speed": speed, **stats(group)})

by_speed_segment = []
for speed in (30, 50, 70):
    for start in (0, 100, 200, 300, 400, 500):
        end = start + 100
        group = [x for x in detail if x["speedGroup"] == speed and x["start"] == start and x["end"] == end]
        by_speed_segment.append({"speed": speed, "interval": f"{start}-{end} m", **stats(group)})
    group = [x for x in detail if x["speedGroup"] == speed and x["start"] >= 600]
    if group:
        by_speed_segment.append({"speed": speed, "interval": "600m后末段", **stats(group)})

payload = {
    "base": str(BASE),
    "sourceA": "v2.2.5.9软件出表：乘 K 平均值",
    "sourceB": "内部软件：乘 K+B",
    "fileCountA": len(files_a), "fileCountB": len(files_b),
    "rowCountA": len(rows_a), "rowCountB": len(rows_b), "matched": len(detail),
    "unmatchedA": len(set(index_a) - set(index_b)), "unmatchedB": len(set(index_b) - set(index_a)),
    "overall": stats(detail), "bySpeed": by_speed, "bySegment": by_segment,
    "stableBySpeed": stable_by_speed, "bySpeedSegment": by_speed_segment, "detail": detail,
}
OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({k: payload[k] for k in ("fileCountA", "fileCountB", "rowCountA", "rowCountB", "matched", "unmatchedA", "unmatchedB")}, ensure_ascii=False))
