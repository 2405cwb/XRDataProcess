import os
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

BASE = Path(os.environ["CUSTOM_KB_BASE"])
OUT = Path(r"D:\job\工作COD\2025\公路\XRDataProcess\output\doc\自定义K_B平整度与真值对比统计.md")
TRUTH = {(200, 300): 1.25, (300, 400): 1.27, (400, 500): 1.43, (500, 600): 1.75}
ORDER = [(200, 300), (300, 400), (400, 500), (500, 600)]


def read_samples():
    samples = []
    books = [p for p in BASE.rglob("*.xlsx") if not p.name.startswith("~$")]
    for path in books:
        group = path.name.split("-", 1)[0]
        project = path.name.split("__", 1)[0]
        book = load_workbook(path, data_only=True, read_only=True)
        sheet = book[book.sheetnames[0]]
        for row in sheet.iter_rows(min_row=4, values_only=True):
            start, end = row[0], row[1]
            if (start, end) not in TRUTH:
                continue
            iri, speed = row[3], row[9]
            if not isinstance(iri, (float, int)):
                raise RuntimeError(f"{path}: {start}-{end} 缺少左IRI")
            samples.append({
                "group": group,
                "project": project,
                "start": int(start), "end": int(end),
                "iri": float(iri), "speed": float(speed), "truth": TRUTH[(start, end)],
            })
    return books, samples


def stats(rows):
    val = np.array([s["iri"] for s in rows])
    truth = np.array([s["truth"] for s in rows])
    err = val - truth
    rel = err / truth
    return {
        "mean": val.mean(), "bias": err.mean(), "mae": abs(err).mean(),
        "rmse": np.sqrt(np.mean(err*err)), "mape": abs(rel).mean(),
        "max_abs_rel": abs(rel).max(), "within5": int(np.sum(abs(rel) <= .05)),
        "positive": int(np.sum(err > 0)), "negative": int(np.sum(err < 0)),
        "min": val.min(), "max": val.max(),
    }


def main():
    books, samples = read_samples()
    if len(books) != 18 or len(samples) != 72:
        raise RuntimeError(f"数据数量异常：workbooks={len(books)} samples={len(samples)}")
    lines=[]; add=lines.append
    add("# 自定义K、B平整度与真值对比统计")
    add("")
    add("## 1. 数据范围")
    add("")
    add("- 输入目录：`平整度加自定义kb`（用户原始描述为“平整度加自定义”）。")
    add("- 输入表：18份 `IRI_100m.xlsx`，30、50、70速度组各6份。")
    add("- 对比样本：每工程取200–300、300–400、400–500、500–600m，共72个100m样本。")
    add("- 表中取值：`左IRI`列；车速取`车速`列。")
    add("- 真值：200–300m=1.25，300–400m=1.27，400–500m=1.43，500–600m=1.75。")
    add("")
    add("偏差定义：`偏差 = 自定义K/B后的IRI − 真值`；`相对误差 = 偏差 / 真值 × 100%`。正值表示偏高，负值表示偏低。")
    add("")
    add("## 2. 总体结论")
    add("")
    add("当前自定义K/B结果存在明显的分速度系统偏差：30、70速度组整体偏低，50速度组整体偏高，且50速度组的偏高最明显。")
    add("")
    add("| 速度组 | 样本数 | IRI均值 | 平均偏差 | MAE | RMSE | 平均相对误差 | 最大相对误差 | ≤5%样本数 | 偏高/偏低样本数 |")
    add("|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
    for group in ("30", "50", "70"):
        st=stats([s for s in samples if s["group"]==group])
        add(f"| {group} | 24 | {st['mean']:.4f} | {st['bias']:+.4f} | {st['mae']:.4f} | {st['rmse']:.4f} | {st['mape']*100:.2f}% | {st['max_abs_rel']*100:.2f}% | {st['within5']}/24 | {st['positive']}/{st['negative']} |")
    allst=stats(samples)
    add(f"| 全部 | 72 | {allst['mean']:.4f} | {allst['bias']:+.4f} | {allst['mae']:.4f} | {allst['rmse']:.4f} | {allst['mape']*100:.2f}% | {allst['max_abs_rel']*100:.2f}% | {allst['within5']}/72 | {allst['positive']}/{allst['negative']} |")
    add("")
    add("解释：30组24个样本全部低于真值；50组24个样本全部高于真值；70组24个样本全部低于真值。该现象不是少数异常点造成，而是当前自定义参数在不同速度档的整体偏置。")
    add("")
    add("## 3. 按真值区间统计")
    add("")
    add("| 速度组 | 区间(m) | 真值 | 自定义IRI均值 | 最小–最大 | 平均偏差 | 平均相对误差 | ≤5%样本数 |")
    add("|---|---|---:|---:|---|---:|---:|---:|")
    for group in ("30", "50", "70"):
        for interval in ORDER:
            rows=[s for s in samples if s["group"]==group and (s["start"],s["end"])==interval]
            st=stats(rows)
            add(f"| {group} | {interval[0]}–{interval[1]} | {rows[0]['truth']:.2f} | {st['mean']:.4f} | {st['min']:.4f}–{st['max']:.4f} | {st['bias']:+.4f} | {st['mape']*100:.2f}% | {st['within5']}/6 |")
    add("")
    add("重点偏差：")
    add("")
    add("- 30组：500–600m平均偏低0.3604 IRI，平均相对误差20.59%。")
    add("- 50组：200–500m三个区间均平均偏高约0.84–0.98 IRI，相对误差约67%–68%。")
    add("- 70组：500–600m平均偏低0.2657 IRI，相对误差15.18%。")
    add("")
    add("## 4. 逐工程详细对比")
    add("")
    add("| 速度组 | 工程 | 区间(m) | 表中左IRI | 车速(km/h) | 真值 | 偏差 | 相对误差 |")
    add("|---|---|---|---:|---:|---:|---:|---:|")
    for group in ("30", "50", "70"):
        rows=sorted([s for s in samples if s["group"]==group], key=lambda s:(s["project"], ORDER.index((s["start"],s["end"]))))
        for s in rows:
            err=s["iri"]-s["truth"]; rel=err/s["truth"]
            add(f"| {group} | {s['project']} | {s['start']}–{s['end']} | {s['iri']:.5f} | {s['speed']:.2f} | {s['truth']:.2f} | {err:+.5f} | {rel*100:+.2f}% |")
    add("")
    add("## 5. 判读")
    add("")
    add("当前这套“自定义K/B”并未在三个速度档同时接近真值：50组存在明显过校正，30与70组仍存在欠校正。若这些参数的目标是让三个速度档都接近同一真值，当前结果不宜作为最终参数。")
    add("")
    add("该统计仅报告表中结果与用户提供真值的偏差，不改变平整度算法，也不推断真值或算法哪一方有问题。")
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines)+"\n", encoding="utf-8")
    print(OUT)


if __name__ == "__main__":
    main()
